import os
import shutil
import openpyxl
from datetime import datetime, date
from typing import List, Dict

REQUIRED_COLUMNS = ["Email", "First Name", "Last Name", "Team Name", "Start Date", "End Date", "Status", "Reason"]

class ExcelManager:
    def __init__(self, file_path: str):
        self.file_path = file_path

    def read_records(self) -> List[Dict[str, str]]:
        if not os.path.exists(self.file_path):
            raise FileNotFoundError(f"Tracker Excel file not found at: {self.file_path}")
        
        wb = openpyxl.load_workbook(self.file_path, data_only=True)
        ws = wb.active
        
        # Read header row (row 1)
        headers = []
        for cell in ws[1]:
            val = str(cell.value).strip() if cell.value is not None else ""
            headers.append(val)
            
        # Clean headers list
        headers = [h for h in headers if h]
        
        # Validate columns
        missing = [col for col in REQUIRED_COLUMNS if col not in headers]
        if missing:
            raise ValueError(f"Excel sheet is missing required columns: {missing}. Found: {headers}")
            
        records = []
        # ws.iter_rows starts at row 2
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            values = [cell.value for cell in row]
            if not any(v is not None for v in values):
                continue # Skip entirely empty rows
                
            row_dict = {}
            for col_name, val in zip(headers, values):
                # Clean up values (strip whitespace and stringify dates/numbers)
                if val is None:
                    row_dict[col_name] = ""
                elif isinstance(val, (datetime, date)):
                    row_dict[col_name] = val.strftime("%Y-%m-%d")
                else:
                    row_dict[col_name] = str(val).strip()
            records.append(row_dict)
            
        wb.close()
        return records

    def create_backup(self) -> str:
        if not os.path.exists(self.file_path):
            return ""
        
        dir_name = os.path.dirname(os.path.abspath(self.file_path))
        backup_dir = os.path.join(dir_name, "backups")
        os.makedirs(backup_dir, exist_ok=True)
        
        base_name = os.path.basename(self.file_path)
        name, ext = os.path.splitext(base_name)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_filename = f"{name}_backup_{timestamp}{ext}"
        backup_path = os.path.join(backup_dir, backup_filename)
        
        shutil.copy2(self.file_path, backup_path)
        return backup_path

    def write_records(self, records: List[Dict[str, str]]) -> str:
        # Create a backup of the existing file first
        backup_path = self.create_backup()
        
        if os.path.exists(self.file_path):
            # Update the existing file in-place to preserve any cell styling or multiple sheets!
            wb = openpyxl.load_workbook(self.file_path)
            ws = wb.active
            
            # Map column names to 1-based column indexes
            header_map = {}
            for col_idx in range(1, ws.max_column + 1):
                val = ws.cell(row=1, column=col_idx).value
                if val is not None:
                    header_map[str(val).strip()] = col_idx
                    
            # For each record, write its fields into the matching cells
            for idx, record in enumerate(records):
                row_idx = idx + 2
                for col_name, val in record.items():
                    if col_name in header_map:
                        ws.cell(row=row_idx, column=header_map[col_name], value=val)
                        
            wb.save(self.file_path)
            wb.close()
        else:
            # Create a brand new workbook if the file doesn't exist
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "GCA Tracker"
            
            # Write headers
            for col_idx, col_name in enumerate(REQUIRED_COLUMNS, start=1):
                ws.cell(row=1, column=col_idx, value=col_name)
                
            # Write data rows
            for row_idx, record in enumerate(records, start=2):
                for col_idx, col_name in enumerate(REQUIRED_COLUMNS, start=1):
                    val = record.get(col_name, "")
                    ws.cell(row=row_idx, column=col_idx, value=val)
                    
            wb.save(self.file_path)
            wb.close()
            
        return backup_path
